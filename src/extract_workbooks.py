"""
extract_workbooks.py — Extract structured data from all 9 Excel workbooks.

No LLM needed — uses openpyxl and pandas directly.
Handles:
  - 6 BOQ workbooks (bill-of-quantity line items)
  - Receivables Ageing workbook
  - Trial Balance workbook
  - Plant & Machinery Register workbook

IMPORTANT: The BRIEFING warns that workbooks contain "live formulas"
and "Notes sheets". We read with data_only=True to get cached values.
"""
import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.config import DOCUMENTS_DIR, EXTRACTED_DIR


WORKBOOKS_DIR = DOCUMENTS_DIR / "workbooks"


def extract_boq_workbook(filepath: Path) -> dict:
    """Extract a Bill of Quantities workbook."""
    # Extract contract number from filename
    # e.g. "BOQ_and_Measurements_Contract_71.xlsx" -> "Contract_71"
    name = filepath.stem
    contract_ref = name.replace("BOQ_and_Measurements_", "")
    
    result = {
        "_doc_id": f"WB-{contract_ref}",
        "_doc_type": "boq_workbook",
        "_source_file": str(filepath),
        "contract_ref": contract_ref,
        "sheets": {},
    }
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Convert all values to serializable types
            clean_row = []
            for cell in row:
                if cell is None:
                    clean_row.append(None)
                elif isinstance(cell, (int, float)):
                    clean_row.append(cell)
                else:
                    clean_row.append(str(cell))
            
            if i == 0:
                headers = clean_row
            else:
                if any(v is not None for v in clean_row):
                    if headers:
                        row_dict = dict(zip(headers, clean_row))
                        rows.append(row_dict)
                    else:
                        rows.append(clean_row)
        
        result["sheets"][sheet_name] = {
            "headers": headers,
            "row_count": len(rows),
            "data": rows,
        }
    
    wb.close()
    return result


def extract_ageing_workbook(filepath: Path) -> dict:
    """Extract the Receivables Ageing workbook."""
    result = {
        "_doc_id": "WB-Receivables_Ageing",
        "_doc_type": "receivables_ageing",
        "_source_file": str(filepath),
        "sheets": {},
    }
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            clean_row = []
            for cell in row:
                if cell is None:
                    clean_row.append(None)
                elif isinstance(cell, (int, float)):
                    clean_row.append(cell)
                else:
                    clean_row.append(str(cell))
            
            if i == 0:
                headers = clean_row
            else:
                if any(v is not None for v in clean_row):
                    if headers:
                        row_dict = dict(zip(headers, clean_row))
                        rows.append(row_dict)
                    else:
                        rows.append(clean_row)
        
        result["sheets"][sheet_name] = {
            "headers": headers,
            "row_count": len(rows),
            "data": rows,
        }
    
    wb.close()
    return result


def extract_trial_balance(filepath: Path) -> dict:
    """Extract the Trial Balance workbook."""
    result = {
        "_doc_id": "WB-Trial_Balance",
        "_doc_type": "trial_balance",
        "_source_file": str(filepath),
        "sheets": {},
    }
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            clean_row = []
            for cell in row:
                if cell is None:
                    clean_row.append(None)
                elif isinstance(cell, (int, float)):
                    clean_row.append(cell)
                else:
                    clean_row.append(str(cell))
            
            if i == 0:
                headers = clean_row
            else:
                if any(v is not None for v in clean_row):
                    if headers:
                        row_dict = dict(zip(headers, clean_row))
                        rows.append(row_dict)
                    else:
                        rows.append(clean_row)
        
        result["sheets"][sheet_name] = {
            "headers": headers,
            "row_count": len(rows),
            "data": rows,
        }
    
    wb.close()
    return result


def extract_asset_register(filepath: Path) -> dict:
    """Extract the Plant & Machinery Register workbook."""
    result = {
        "_doc_id": "WB-Plant_Machinery",
        "_doc_type": "asset_register",
        "_source_file": str(filepath),
        "sheets": {},
    }
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            clean_row = []
            for cell in row:
                if cell is None:
                    clean_row.append(None)
                elif isinstance(cell, (int, float)):
                    clean_row.append(cell)
                else:
                    clean_row.append(str(cell))
            
            if i == 0:
                headers = clean_row
            else:
                if any(v is not None for v in clean_row):
                    if headers:
                        row_dict = dict(zip(headers, clean_row))
                        rows.append(row_dict)
                    else:
                        rows.append(clean_row)
        
        result["sheets"][sheet_name] = {
            "headers": headers,
            "row_count": len(rows),
            "data": rows,
        }
    
    wb.close()
    return result


def extract_all_workbooks():
    """Extract all 9 workbooks in the documents/workbooks directory."""
    if not WORKBOOKS_DIR.exists():
        print(f"Workbooks directory not found: {WORKBOOKS_DIR}")
        return
    
    files = list(WORKBOOKS_DIR.glob("*.xlsx"))
    print(f"Found {len(files)} workbooks to extract")
    
    for filepath in files:
        print(f"  Extracting: {filepath.name}")
        
        # Determine workbook type and extract accordingly
        name = filepath.stem.lower()
        
        if name.startswith("boq"):
            data = extract_boq_workbook(filepath)
        elif "ageing" in name or "receivable" in name:
            data = extract_ageing_workbook(filepath)
        elif "trial" in name and "balance" in name:
            data = extract_trial_balance(filepath)
        elif "plant" in name or "machinery" in name or "asset" in name:
            data = extract_asset_register(filepath)
        else:
            # Generic extraction
            data = extract_boq_workbook(filepath)
            data["_doc_id"] = f"WB-{filepath.stem}"
            data["_doc_type"] = "unknown_workbook"
        
        # Save to extracted directory
        output_path = EXTRACTED_DIR / f"{data['_doc_id']}.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"    -> Saved to {output_path.name}")
    
    print(f"\nAll workbooks extracted!")


if __name__ == "__main__":
    extract_all_workbooks()
