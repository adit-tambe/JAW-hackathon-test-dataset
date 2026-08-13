"""
discover.py — find every document under a root and decide what each one is,
using only its contents.

At run time we are handed a directory we have never seen. The tree is nested by
document type, but the nesting will not match anything in our samples, and the
file names are not ours. So nothing in this module may key off a path, a folder
name or a file name: a document is typed by what it says on its face.

The previous ingestion path read a checked-in manifest (`document_index.csv`)
that mapped 687 known file names to types, then resolved each against a fixed
`documents/` directory. That works only on the machine the manifest was built
on. Everywhere else it silently finds nothing, which is the worst possible
failure: an empty database and a full set of confident, wrong answers.

Doc ids are assigned here rather than read from anywhere. Downstream loaders
select extracted records by id prefix, so the prefixes below are a contract with
`build_db.py` and must not be renamed casually. Nothing downstream parses
meaning out of an id — fiscal years and the like are read from document text —
so renumbering is safe.
"""
from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

try:
    import fitz  # PyMuPDF
except ImportError:  # pragma: no cover - surfaced by run.sh preflight
    fitz = None
try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


# ── Type → id prefix ────────────────────────────────────────────────────────
# The prefix is what build_db.py globs on.
PREFIX = {
    "completion_certificate":         "DOC-CC",
    "company_completion_certificate": "DOC-CCC",
    "reference_letter":               "DOC-REF",
    "personnel_certificate":          "DOC-PCERT",
    "cv":                             "DOC-CV",
    "performance_bond":               "DOC-BOND",
    "financial_statement":            "DOC-FS",
    "general_ledger_book":            "DOC-GLB",
    "bank_statement":                 "DOC-BANK",
    "iso_certificate":                "DOC-CERT",
    "ra_bill":                        "DOC-RABILL",
    "final_ra_bill":                  "DOC-FINBILL",
    "tender_dossier":                 "DOC-DOSSIER",
    "compliance_matrix":              "DOC-CM",
    "annual_report":                  "DOC-AR",
    "past_performance_portfolio":     "DOC-PPP",
    # workbooks all share the WB- prefix; the specific type rides in _doc_type
    "ageing_workbook":                "WB-AGEING",
    "trial_balance_workbook":         "WB-TB",
    "asset_register_workbook":        "WB-ASSETS",
    "boq_workbook":                   "WB-BOQ",
    "unknown_workbook":               "WB-OTHER",
}


# ── PDF typing ──────────────────────────────────────────────────────────────
# Ordered most-specific first; the first rule whose markers all appear wins.
# Markers are matched against whitespace-collapsed lowercase text, because the
# PDFs render several headings in small-caps ("ComplianCe CheCklist") and wrap
# titles across lines.
PDF_RULES: list[tuple[str, tuple[str, ...]]] = [
    # Two certificates share most of their vocabulary, and each comes in two
    # layouts. The company's own record says either "record of work completed"
    # or "completion certificate issued by the contractor"; the client-issued
    # one says "work completion certificate". Test the contractor's own first,
    # because its second layout also contains the words "completion
    # certificate" and would otherwise be filed as the client's.
    ("company_completion_certificate", ("record of work completed",)),
    ("company_completion_certificate", ("completion certificate issued by the contractor",)),
    ("completion_certificate",         ("work completion certificate",)),
    ("completion_certificate",         ("certificate of completion",)),

    ("reference_letter",               ("letter of recommendation",)),
    ("reference_letter",               ("to whomsoever it may concern",)),
    ("reference_letter",               ("subject: performance of",)),
    ("past_performance_portfolio",     ("portfolio of completed works",)),
    ("past_performance_portfolio",     ("past performance portfolio",)),

    # "final running account bill" contains "running account bill", so the
    # final variant must be tested first or every final bill reads as interim.
    ("final_ra_bill",                  ("final running account bill",)),
    ("final_ra_bill",                  ("final bill with measurement",)),
    ("ra_bill",                        ("running account bill",)),

    ("performance_bond",               ("performance bank guarantee",)),
    ("performance_bond",               ("subject: performance bond",)),
    ("iso_certificate",                ("certificate of registration",)),
    ("cv",                             ("curriculum vitae",)),
    ("personnel_certificate",          ("credential id",)),
    ("personnel_certificate",          ("this credential is conferred upon",)),
    ("tender_dossier",                 ("tender submission dossier",)),
    ("compliance_matrix",              ("compliance checklist",)),
    ("compliance_matrix",              ("bid compliance",)),
    ("annual_report",                  ("annual report", "corporate information")),
    ("financial_statement",            ("statement of profit and loss",)),
    ("financial_statement",            ("audited financial results",)),
    ("general_ledger_book",            ("general ledger",)),
    ("bank_statement",                 ("statement of account",)),
    ("bank_statement",                 ("account statement",)),
]

# Fallback scoring, used only when no rule fires. Weaker, field-level evidence:
# a document that never announces itself can still be recognised by the shape of
# the fields it carries.
PDF_HINTS: dict[str, tuple[str, ...]] = {
    "completion_certificate":  ("particulars of the work", "general conditions of contract"),
    "reference_letter":        ("this office has engaged", "recommendation"),
    "cv":                      ("employee id", "total experience", "qualification"),
    "personnel_certificate":   ("issuing authority", "credential type"),
    "performance_bond":        ("irrevocable", "bond no"),
    "bank_statement":          ("withdrawal", "deposit", "opening balance"),
    "general_ledger_book":     ("chart of accounts", "posted lines"),
    "financial_statement":     ("profit and loss", "current year", "previous year"),
    "compliance_matrix":       ("requirement", "complied", "evidence"),
    "tender_dossier":          ("tender inviting authority", "bid value"),
    "ra_bill":                 ("abstract of work done", "bill no"),
    "annual_report":           ("registered office", "cin"),
    "iso_certificate":         ("management system", "certification body"),
}

_WS = re.compile(r"\s+")


def normalise(text: str) -> str:
    """Lowercase and collapse whitespace so multi-word markers survive wrapping."""
    return _WS.sub(" ", text.lower())


def sniff_pdf_type(text: str) -> tuple[str | None, str]:
    """Return (doc_type, how_it_was_decided) for a PDF's extracted text."""
    head = normalise(text[:4000])
    for doc_type, markers in PDF_RULES:
        if all(m in head for m in markers):
            return doc_type, f"marker:{markers[0]}"

    # Nothing announced itself. Score the whole document on field-level hints.
    whole = normalise(text[:20000])
    best, best_score = None, 0
    for doc_type, hints in PDF_HINTS.items():
        score = sum(1 for h in hints if h in whole)
        if score > best_score:
            best, best_score = doc_type, score
    if best and best_score >= 2:
        return best, f"hints:{best_score}"
    return None, "unmatched"


# ── Workbook typing ─────────────────────────────────────────────────────────
def sniff_workbook_type(sheet_names: list[str],
                        headers: dict[str, list[str]]) -> tuple[str | None, str]:
    """Type a workbook from its sheet names and header rows."""
    names = [s.lower().strip() for s in sheet_names]
    flat = " | ".join(
        " ".join(str(h).lower() for h in (headers.get(s) or []) if h)
        for s in sheet_names
    )

    if any(n.startswith("ar ageing") or "ageing" in n for n in names) or (
            "invoiced" in flat and "outstanding" in flat):
        return "ageing_workbook", "ar-ageing"
    if any("plant register" in n or "asset" in n for n in names) or "asset id" in flat:
        return "asset_register_workbook", "asset-register"
    if any(n.startswith("tb ") or "trial balance" in n for n in names) or (
            "debit" in flat and "credit" in flat and "account" in flat):
        return "trial_balance_workbook", "trial-balance"
    if any(n == "boq" for n in names) or ("item no" in flat and "rate" in flat):
        return "boq_workbook", "boq"
    return "unknown_workbook", "unmatched"


_CONTRACT_RE = re.compile(r"contract\s*#?\s*(\d+)", re.I)


def workbook_contract_ref(headers_and_notes: str, fallback: str) -> str:
    """BOQ workbooks name their contract in the Notes sheet, not the file name."""
    m = _CONTRACT_RE.search(headers_and_notes)
    return f"Contract_{m.group(1)}" if m else fallback


# ── Walking ─────────────────────────────────────────────────────────────────
@dataclass
class Doc:
    doc_id: str
    doc_type: str
    path: Path
    kind: str          # "pdf" | "xlsx"
    text: str = ""     # PDFs only; kept so callers need not re-open the file
    detail: str = ""   # how the type was decided, for the run log


PDF_SUFFIXES = {".pdf"}
XLSX_SUFFIXES = {".xlsx", ".xlsm"}


def _read_pdf_text(path: Path) -> str:
    doc = fitz.open(path)
    try:
        return "\n".join(page.get_text() for page in doc)
    finally:
        doc.close()


def _read_workbook_shape(path: Path) -> tuple[list[str], dict[str, list], str]:
    """Sheet names, header row per sheet, and a blob of note text."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        names = list(wb.sheetnames)
        headers: dict[str, list] = {}
        notes: list[str] = []
        for name in names:
            ws = wb[name]
            first = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    first = [c for c in row]
                if name.lower().startswith("note"):
                    notes.extend(str(c) for c in row if c)
                if i >= 8:
                    break
            headers[name] = first or []
        return names, headers, " ".join(notes)
    finally:
        wb.close()


def discover(root: Path, verbose: bool = True) -> list[Doc]:
    """Walk `root` recursively and type every PDF and workbook beneath it."""
    root = Path(root)
    if not root.exists():
        raise FileNotFoundError(f"--docs path does not exist: {root}")

    files = [p for p in sorted(root.rglob("*"))
             if p.is_file() and p.suffix.lower() in (PDF_SUFFIXES | XLSX_SUFFIXES)]
    if verbose:
        print(f"  Walking {root}")
        print(f"  Found {len(files)} candidate files "
              f"({sum(1 for f in files if f.suffix.lower() in PDF_SUFFIXES)} pdf, "
              f"{sum(1 for f in files if f.suffix.lower() in XLSX_SUFFIXES)} xlsx)")

    counters: dict[str, int] = {}
    docs: list[Doc] = []
    unmatched: list[Path] = []

    for path in files:
        try:
            if path.suffix.lower() in PDF_SUFFIXES:
                if fitz is None:
                    raise RuntimeError("PyMuPDF (fitz) is not installed")
                text = _read_pdf_text(path)
                doc_type, how = sniff_pdf_type(text)
                kind = "pdf"
            else:
                if openpyxl is None:
                    raise RuntimeError("openpyxl is not installed")
                names, headers, notes = _read_workbook_shape(path)
                doc_type, how = sniff_workbook_type(names, headers)
                text = notes
                kind = "xlsx"
        except Exception as exc:                      # a single bad file must not
            print(f"  !! could not read {path.name}: {exc}")   # sink the whole run
            unmatched.append(path)
            continue

        if doc_type is None:
            unmatched.append(path)
            continue

        n = counters.get(doc_type, 0) + 1
        counters[doc_type] = n
        docs.append(Doc(doc_id=f"{PREFIX[doc_type]}-{n:05d}", doc_type=doc_type,
                        path=path, kind=kind, text=text, detail=how))

    if verbose:
        print(f"  Typed {len(docs)} documents:")
        for doc_type in sorted(counters, key=lambda t: -counters[t]):
            print(f"      {doc_type:34s} {counters[doc_type]}")
        if unmatched:
            print(f"  {len(unmatched)} file(s) could not be typed:")
            for p in unmatched[:10]:
                print(f"      {p}")
    return docs
